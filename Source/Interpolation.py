#from tqdm import tqdm
import numpy as np
import pandas as pd
import openpyxl
import string
#from PyQt5.QtWidgets import QProgressBar

# =============================================================================
# # required inpuuts 
# fp = 'C:/Users/rjw3/OneDrive/Desktop/Test.xlsx'
# sn = 'Sheet1'
# steps =.001
# col_to_interp_from = None
# 
# 
# #### check for optional user input
# fp2 = str()
# sn2 = str()
# same_sheet = True
# same_file = True
# =============================================================================


def interp(x1,y1, x3,y3, x2):
    top = (x2 - x1) * (y3 - y1)
    bot = (x3 - x1)
    y2 = (top/bot) + y1
    return y2

#inatlize
first_row = int()
first_col = str()

#### add col to interplolate

#return true if input is NaN
def isNaN(num):
    return (num!= num) or (num == None)

#deleta all NaN
def find_table(DataFrame):
    #check for empty rows
    row_found = False
    col_found = False
    for i in (DataFrame.index.values):
        test_arr = np.array([])
        for o in DataFrame.columns.values:
            test_arr = np.append(test_arr,
                                 isNaN(DataFrame.loc[i, o]))
        if(np.sum(test_arr) == np.size(test_arr)):
            DataFrame = DataFrame.drop(i)
        else:
            if(row_found==False):
                first_row = i + 1 
                #print('\n' + 'first row =' + str(first_row))
                row_found = True
    #check for empty columns
    for i in (DataFrame.columns.values):
        test_arr = np.array([])
        for o in DataFrame.index.values:
            test_arr = np.append(test_arr,
                                 isNaN(DataFrame.loc[o, i]))
        if(np.sum(test_arr) == np.size(test_arr)):
            DataFrame = DataFrame.drop(i, axis=1)
        else:
            if(col_found==False):
                first_col = i
                col_found = True
    found=True
    try:
        return DataFrame, first_row, first_col, found
    
    except UnboundLocalError:
        found=False
        first_row = None
        first_col = None
        return DataFrame, first_row, first_col, found

#set anything that is not a float or int to NaN
def NaN_non_values(Dataframe):
    for i in (Dataframe.columns.values):
        for o in Dataframe.index.values:
            y = Dataframe.loc[o,i]
            if((isinstance(y,float) or isinstance(y, int)) != True):
                Dataframe.loc[o,i] = None
    return Dataframe



#create dictionary for excel columns
alph_dic = dict()
L = np.arange(0,25)
numb = 0
st = str()
while(numb<10000):
    for i in range(0,25):
        for o in L:
            alph_dic[numb] = str(st) + string.ascii_uppercase[o]
            numb = numb + 1
        st = str(st) + str(string.ascii_uppercase[i])


# rename columns with the appropriate excel keys
def rename_col(DataFrame):
    dic = dict()
    for i in DataFrame.columns.values:
        key = alph_dic[i]
        dic[i] = key
    DataFrame = DataFrame.rename(columns=dic)
    return DataFrame


def int_table(fp, sn, steps, col_to_interp_from, fp2, sn2, same_sheet, 
              same_file, QProgressBar, QLabel):

    #pull initial table
    try:
        table = pd.read_excel(fp, sheet_name=sn, header=None)
    except PermissionError:
            return 'Close the file & try again'
    table = rename_col(table)
    table = NaN_non_values(table)
    table, first_row, first_col, Data_found = find_table(table)
    if(Data_found==False):
        return 'No data Found'
    # interpolate and create new table
    n_table = pd.DataFrame(columns=table.columns.values)
    current_row = first_row
    if((col_to_interp_from==None) or(col_to_interp_from=='')):
        col = first_col
    else:
        col = col_to_interp_from
    
    #check for appropriate step size
    bol_test = True
    
    QProgressBar.setMaximum(len(table.index.values))
    progress = 0
    for i in (table.index.values):
        if(i < len(table.index.values)):
            test1 = table.loc[i+1, col] > (table.loc[i, col] + steps)
        if(test1==False):
            bol_test = False
    if(bol_test==False):
        return 'Inappropriate step size'
    #interpolate
    for i in (table.index.values):
        progress = progress + 1
        QProgressBar.setValue(progress)
        x1 = table.loc[i,col]
        x2 = x1 + steps
        if(i == first_row-1):
            x3 = table.loc[i + 1,col]
            n_table.loc[current_row, :] = table.loc[i, :]
        else:
            
            if(i < ((len(table.index.values)-2) + first_row)):
                current_row = current_row + 1
                x3 = table.loc[i + 1,col]
                n_table.loc[current_row, :] = table.loc[i, :]
            else:
                x3 = x2
                current_row = current_row + 1
                n_table.loc[current_row, :] = table.loc[i, :]
        while(round(x2 - x3,5) < 0):
            current_row = current_row + 1
            n_table.loc[current_row , col] = x2
            for o in table.columns.values:
                if(o!=col):
                    y1 = table.loc[i, o]
                    y3 = table.loc[i + 1, o]
                    y2 = interp(x1,y1, x3,y3, x2)
                    n_table.loc[current_row, o] = y2

            x2 = x2 + steps
    
    QLabel.setText('Saving Table')
    QLabel.move(10, 875)
    QLabel.show()
    if(same_file==True):
        wb = openpyxl.load_workbook(fp)
        if(same_sheet==True):
            work_sheet = wb[sn]
            for o in n_table.index.values:
                row = str(o)
                for i in n_table.columns.values:
                    col = str(i)
                    cell = col + row
                    work_sheet[cell] = n_table.loc[o, i]

            wb.save(fp)
            
        else:
            try:
                work_sheet = wb[sn2]
            except KeyError:
                wb.create_sheet(sn2)
                work_sheet = wb[sn2]
            for o in n_table.index.values:
                row = str(o)
                for i in n_table.columns.values:
                    col = str(i)
                    cell = col + row
                    work_sheet[cell] = n_table.loc[o, i]

            wb.save(fp)
            
    else:
        try:
            wb = openpyxl.load_workbook(fp2)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        if(same_sheet==True):
            wb.create_sheet(sn)
            work_sheet = wb[sn]
            for o in n_table.index.values:
                row = str(o)
                for i in n_table.columns.values:
                    col = str(i)
                    cell = col + row
                    work_sheet[cell] = n_table.loc[o, i]

            wb.save(fp2)
            
        else:
            try:
                work_sheet = wb[sn2]
            except KeyError:
                wb.create_sheet(sn2)
                work_sheet = wb[sn2]
            for o in n_table.index.values:
                row = str(o)
                for i in n_table.columns.values:
                    col = str(i)
                    cell = col + row
                    work_sheet[cell] = n_table.loc[o, i]
            wb.save(fp2)
    
   
    return 'Interpolation complete'

